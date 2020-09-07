import sys
import pyperclip
import openpyxl
import configparser

def read_ini():
    #iniファイルの読み込み
   inifile = configparser.ConfigParser()
    #raw文字列を使ってパスの指定
   inifile.read(r'C:\Users\kinsh\PycharmProjects\excelpro\pw.ini')
   #変数値の取得
   local_pw_file = inifile.get('ENV', 'LOCAL_PW_FILE')
   return(local_pw_file)

def get_pw(category_name, local_pw_dir):
    '''カテゴリ名を引数として受け取り、Excelに保存されている対象パスワードを取得する'''

    # パスワード管理ファイルのオープン
    try:
        wb = openpyxl.load_workbook(local_pw_file)
    except OSError:
        print('【ERROR】: 指定されたファイル[LOCAL_PW_FILE]が存在しません。', local_pw_file)
        sys.exit()

    # Excelの「パスワード一覧」シートをアクティブ化
    sheet = wb['パスワード一覧']
    # 有効セルの列数の取得
    col_num = sheet.max_column
    # 有効セルの行数の取得
    row_num = sheet.max_row
    # 最終セル位置の取得
    end_cell = 'B' + str(row_num)

    # パスワードファイルの全情報をタプル化し抽出
    pw_info = tuple(sheet['A1':end_cell])
    # パスワード取得有無のフラグ
    flag = False
    password = ''
    # タプル化したpw_infoから該当のカテゴリ名が記録されているセルを検索、パスワード情報を取得
    for i in range(row_num): # 行方向のループ処理
        for j in range(col_num):  # 列方向のループ処理
            if pw_info[i][j].value == category_name:  # カテゴリ名が見つかったら
                password = pw_info[i][j + 1].value  # 該当行のパスワード列の値を取得
                flag = True  # パスワードが見つかったらflagをTrueにセット
                break
            else:
                break
        if flag: # flag=Trueだったらループを抜ける。
            break

# パスワード情報を戻す
    return password

local_pw_file = read_ini()
if len(sys.argv) < 2 or len(sys.argv) > 2:
   print('使い方: pw [カテゴリ名]')
   print('第1引数にカテゴリ名(必須）を指定します。')
   sys.exit()
elif len(sys.argv) == 2:
   category_name = sys.argv[1]
   passwords = get_pw(category_name, local_pw_file)
   if len(passwords) == 0:
       print('存在しないか、誤ったカテゴリ名を指定しています。')
       print('入力値を確認してください')
       sys.exit()
   pyperclip.copy(passwords)
   print('【カテゴリ名:' + category_name + '】のパスワードをクリップボードにコピーしました。')
   sys.exit()